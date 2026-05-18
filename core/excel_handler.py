"""
우편모아 / 라벨텍 / 도청방문등록 엑셀 생성 모듈
"""
from pathlib import Path
from datetime import date
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

OUTPUT_DIR = Path(__file__).parent.parent / "output"
HEARING_OUTPUT_DIR = Path(__file__).parent.parent / "output" / "위원회개최"
RESULT_OUTPUT_DIR  = Path(__file__).parent.parent / "output" / "위원회결과보고"
RESULT_SUSANG_TEMPLATE = Path(__file__).parent.parent / "템플릿" / "위원회 결과보고" / "수당 지급내역.xlsx"


def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def generate_woopyeonmoa(cases: list[dict]) -> Path:
    """
    우편모아 포맷 xlsx 생성 (원본 xls 컬럼 순서 그대로)
    수수료*=보통, 환부*=환부, 규격*=규격외, 중량=25 고정
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    fname = f"우편모아_{date.today().strftime('%Y%m%d')}.xlsx"
    out_path = OUTPUT_DIR / fname

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "우편모아"

    headers = [
        "수수료*", "환부*", "규격*", "중량",
        "수취인*", "우편번호*", "기본주소*", "상세주소",
        "휴대폰", "문서번호", "문서제목", "비고",
    ]

    hdr_fill = PatternFill("solid", fgColor="FFFF00")
    hdr_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()

    current_row = 2
    for case in cases:
        def _wp_row(성명, 우편번호, 주소, 연락처, 접수번호):
            return [
                "보통", "환부", "규격외", 25,
                성명, 우편번호, 주소, "", 연락처, 접수번호, "집합건물 분쟁조정 통지", "",
            ]

        for row_data in [
            _wp_row(case.get("피신청인_성명", ""), case.get("피신청인_우편번호", ""),
                    case.get("피신청인_주소", ""), case.get("피신청인_연락처", ""),
                    case.get("접수번호", "")),
            *(
                [_wp_row(case.get("피신청인2_성명", ""), case.get("피신청인2_우편번호", ""),
                         case.get("피신청인2_주소", ""), case.get("피신청인2_연락처", ""),
                         case.get("접수번호", ""))]
                if case.get("피신청인2_성명") else []
            ),
        ]:
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.border = _thin_border()
                cell.alignment = Alignment(vertical="center")
            current_row += 1

    widths = [10, 8, 8, 6, 16, 10, 40, 20, 14, 12, 18, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(out_path)
    return out_path


def generate_labeltek(cases: list[dict]) -> Path:
    """
    라벨텍 포맷 xlsx 생성
    - 제목 열: "집합건물 분쟁조정 관련" (고정)
    - 이후: 주소, 이름, 우편번호
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    fname = f"라벨텍_{date.today().strftime('%Y%m%d')}.xlsx"
    out_path = OUTPUT_DIR / fname

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "라벨텍"

    headers = ["제목", "주소", "이름", "우편번호"]
    hdr_fill = PatternFill("solid", fgColor="D9E1F2")
    hdr_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin_border()

    current_row = 2
    for case in cases:
        def _lb_row(주소, 성명, 우편번호):
            return ["집합건물 분쟁조정 관련", 주소, 성명, 우편번호]

        for row_data in [
            _lb_row(case.get("피신청인_주소", ""), case.get("피신청인_성명", ""),
                    case.get("피신청인_우편번호", "")),
            *(
                [_lb_row(case.get("피신청인2_주소", ""), case.get("피신청인2_성명", ""),
                          case.get("피신청인2_우편번호", ""))]
                if case.get("피신청인2_성명") else []
            ),
        ]:
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=val)
                cell.border = _thin_border()
                cell.alignment = Alignment(vertical="center")
            current_row += 1

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10

    wb.save(out_path)
    return out_path


def _find_docheong_template() -> Path:
    """도청방문등록_양식.xls 템플릿 경로 찾기"""
    template_dir = Path(__file__).parent.parent / "템플릿" / "위원회 개최"
    for f in template_dir.iterdir():
        if f.name.encode("utf-8").startswith(b"\xeb\x8f\x84\xec\xb2\xad"):
            return f
    return template_dir / "도청방문등록_양식.xls"


def generate_docheong_visit(case: dict, hearing: dict, members: list[dict]) -> Path:
    """도청방문등록 xlsx 생성 — 경기도 방문등록 양식 (성명·회사명·직책·휴대전화번호·그룹)"""
    접수번호 = case.get("접수번호", "unknown")
    회차 = hearing.get("회차", 1)

    out_dir = HEARING_OUTPUT_DIR / f"{접수번호}_{회차}차"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"도청방문등록_{접수번호}_{회차}차.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "방문등록"

    # ── 헤더 (첫 번째 이미지 양식 그대로)
    headers = [
        "(필수) 성명",
        "(필수) 회사명",
        "직책",
        "(필수) 휴대전화번호\n(ex. 010-1234-5678)",
        "그룹 (ex. 1 ~ 9)\n※ 같은 그룹 선택시 첫번째 인솔자,\n이후 동행인으로 등록합니다.",
    ]
    hdr_fill = PatternFill("solid", fgColor="FFFF00")
    hdr_font = Font(bold=True)
    b = _thin_border()
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = b
    ws.row_dimensions[1].height = 50
    for i, w in enumerate([14, 30, 16, 24, 42], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    def _add_row(r, 성명, 회사명, 직책, 핸드폰):
        for col, val in enumerate([성명, 회사명, 직책, 핸드폰, ""], 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = b
            cell.alignment = Alignment(vertical="center")

    # ── 데이터 행
    role_order = {"위원장": 0, "위원": 1, "간사": 2}
    sorted_members = sorted(members, key=lambda m: role_order.get(m.get("역할", "위원"), 1))

    row = 2
    for m in sorted_members:
        _add_row(row,
                 m.get("성명", ""),
                 m.get("소속", "") or "",
                 m.get("직위", "") or "",
                 m.get("핸드폰번호", "") or "")
        row += 1

    _add_row(row,
             case.get("신청인_성명", ""),
             "",
             case.get("신청인_지위", "") or "신청인",
             case.get("신청인_연락처", "") or "")
    row += 1

    _add_row(row,
             case.get("피신청인_성명", ""),
             "",
             case.get("피신청인_지위", "") or "피신청인",
             case.get("피신청인_연락처", "") or "")

    wb.save(str(out_path))
    return out_path


# ────────────────────────────────────────────────
# 참석수당 지급내역 엑셀 생성
# ────────────────────────────────────────────────

SUSANG_TEMPLATE = Path(__file__).parent.parent / "템플릿" / "위원회 개최" / "참석수당지급내역_양식.xlsx"


def generate_susang_excel(case: dict, hearing: dict, members: list[dict]) -> Path:
    """
    참석수당 지급내역 xlsx 생성.
    템플릿(참석수당지급내역_양식.xlsx)을 로드해 헤더 변수 치환 후
    위원 데이터 행을 덮어씀.
    """
    from datetime import datetime as _dt

    접수번호 = case.get("접수번호", "unknown")
    회차     = hearing.get("회차", 1)

    out_dir = HEARING_OUTPUT_DIR / f"{접수번호}_{회차}차"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"수당지급내역_{접수번호}_{회차}차.xlsx"

    wb = openpyxl.load_workbook(str(SUSANG_TEMPLATE))
    ws = wb.active

    # ── 헤더 변수 치환 (A3, A4)
    def _fmt_dt(dt_str):
        if not dt_str:
            return ""
        try:
            dt = _dt.fromisoformat(str(dt_str).replace(" ", "T"))
            ampm = "오후" if dt.hour >= 12 else "오전"
            h12  = dt.hour - 12 if dt.hour > 12 else (12 if dt.hour == 0 else dt.hour)
            return f"{dt.year}. {dt.month}. {dt.day}.({ampm} {h12}시)"
        except Exception:
            return str(dt_str)

    일시   = _fmt_dt(hearing.get("개최예정일시"))
    장소   = hearing.get("개최장소", "") or ""
    내용   = case.get("신청내용", "") or ""
    접수연도 = case.get("접수연도", date.today().year)

    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str):
                v = v.replace("{{일시 }}", 일시).replace("{{일시}}", 일시)
                v = v.replace("{{장소}}", 장소)
                cell.value = v

    # A5: 조정내용, A6: 참석위원 수
    if ws["A5"].value and "조정내용" in str(ws["A5"].value):
        ws["A5"].value = f"O 조정내용 : {내용}"

    role_order = {"위원장": 0, "위원": 1, "간사": 2}
    sorted_mems = sorted(members, key=lambda m: role_order.get(m.get("역할", "위원"), 1))
    n = len(sorted_mems)

    if ws["A6"].value and "참석위원" in str(ws["A6"].value):
        ws["A6"].value = f"O 참석위원 : {n}명"

    # A10 (계 행) B10: 인원수
    ws["B10"].value = f"{n}명"

    # ── 데이터 행 (행 11~): 기존 3행 분량을 지우고 실제 위원 수만큼 채움
    DATA_START = 11
    DATA_COLS  = list(range(1, 10))  # A~I

    # 기존 데이터 행 초기화 (최대 7행)
    for r in range(DATA_START, DATA_START + 7):
        for c in DATA_COLS:
            ws.cell(row=r, column=c).value = None

    # 위원 데이터 채우기
    b = _thin_border()
    al_c = Alignment(horizontal="center", vertical="center")
    al_l = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    for idx, m in enumerate(sorted_mems):
        r = DATA_START + idx
        birth = str(m.get("생년월일") or "").replace("-", "")[:8]

        ws.cell(row=r, column=1).value = idx + 1         # 연번
        ws.cell(row=r, column=2).value = m.get("성명", "")   # 성명
        ws.cell(row=r, column=3).value = birth            # 생년월일
        ws.cell(row=r, column=4).value = None             # 사전검토수당 (공란)
        ws.cell(row=r, column=5).value = 200000           # 참석수당 기본
        ws.cell(row=r, column=6).value = None             # 참석수당 추가
        ws.cell(row=r, column=7).value = (                # 지급총액 (수식)
            f"=IFERROR(D{r},0)+E{r}+IFERROR(F{r},0)"
        )
        ws.cell(row=r, column=8).value = m.get("은행명", "") or ""   # 은행
        ws.cell(row=r, column=9).value = m.get("계좌번호", "") or ""  # 계좌

        for c in DATA_COLS:
            ws.cell(row=r, column=c).border    = b
            ws.cell(row=r, column=c).alignment = al_c if c != 9 else al_l

    # 계 행 합계 수식 업데이트
    end_row = DATA_START + n - 1
    ws["D10"].value = f"=IFERROR(SUM(D{DATA_START}:D{end_row}),0)"
    ws["E10"].value = f"=SUM(E{DATA_START}:E{end_row})"
    ws["F10"].value = f"=IFERROR(SUM(F{DATA_START}:F{end_row}),0)"
    ws["G10"].value = f"=SUM(G{DATA_START}:G{end_row})"

    # 비고 행 이동 (기존 A14 → 실제 DATA_START+n+1 행)
    note_row = DATA_START + n + 1
    ws.cell(row=note_row, column=1).value = "※ 참석수당 : 2시간까지 200,000원(2시간 초과 100,000원)"

    wb.save(str(out_path))
    return out_path


# ────────────────────────────────────────────────
# 위원회 결과보고 수당 지급내역 (템플릿 서식 보존)
# ────────────────────────────────────────────────

def generate_result_susang_excel(case: dict, hearing: dict, members: list[dict]) -> Path:
    """
    템플릿(위원회 결과보고/수당 지급내역.xlsx)의 서식을 그대로 유지하면서
    일시·장소·조정내용·성명·생년월일·은행·계좌번호만 채워 넣음.
    """
    from datetime import datetime as _dt

    접수번호  = case.get("접수번호", "unknown")
    회차      = hearing.get("회차", 1)
    접수연도  = case.get("접수연도", date.today().year)

    out_dir = HEARING_OUTPUT_DIR / f"{접수번호}_{회차}차" / "결과보고"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"수당지급내역_{접수번호}_{회차}차.xlsx"

    wb = openpyxl.load_workbook(str(RESULT_SUSANG_TEMPLATE))
    ws = wb.active

    def _fmt_dt(dt_str):
        if not dt_str:
            return ""
        try:
            dt = _dt.fromisoformat(str(dt_str).replace(" ", "T"))
            요일 = ["월", "화", "수", "목", "금", "토", "일"][dt.weekday()]
            return f"{dt.year}. {dt.month}. {dt.day}.({요일}) {dt.hour:02d}:{dt.minute:02d}"
        except Exception:
            return str(dt_str)

    일시  = _fmt_dt(hearing.get("개최예정일시"))
    장소  = hearing.get("개최장소", "") or ""
    내용  = hearing.get("조정내용", "") or case.get("신청내용", "") or ""

    role_order  = {"위원장": 0, "위원": 1, "간사": 2}
    sorted_mems = sorted(members, key=lambda m: role_order.get(m.get("역할", "위원"), 1))
    n = len(sorted_mems)

    # 제목 갱신 (A1)
    ws["A1"].value = f"{접수연도}년 제{회차}회 경기도 집합건물 분쟁조정위원회 회의 수당지급내역"

    # 정보 행 — "O 레이블 : 값" 형식 그대로 유지
    ws["A3"].value = f"O 일      시 : {일시}"
    ws["A4"].value = f"O 장      소 : {장소}"
    ws["A5"].value = f"O 조정내용 : {내용}"
    ws["A6"].value = f"O 참석위원 : {n}명"
    ws["B10"].value = f"{n}명"

    # 위원 데이터 행 (11행~): 성명/생년월일/은행/계좌만 덮어씀, 나머지 서식 유지
    DATA_START = 11
    MAX_ROWS   = 7

    for i in range(MAX_ROWS):
        row = DATA_START + i
        if i < n:
            m = sorted_mems[i]
            birth_raw = str(m.get("생년월일") or "").replace("-", "")
            if len(birth_raw) == 8:
                birth_val = int(birth_raw[2:])
            elif len(birth_raw) == 6:
                birth_val = int(birth_raw)
            else:
                birth_val = None
            ws.cell(row=row, column=1).value = i + 1
            ws.cell(row=row, column=2).value = m.get("성명", "")
            ws.cell(row=row, column=3).value = birth_val
            ws.cell(row=row, column=8).value = m.get("은행명", "") or ""
            ws.cell(row=row, column=9).value = m.get("계좌번호", "") or ""
        else:
            for col in (1, 2, 3, 8, 9):
                ws.cell(row=row, column=col).value = None

    wb.save(str(out_path))
    return out_path
