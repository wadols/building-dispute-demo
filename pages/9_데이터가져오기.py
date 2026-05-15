"""
데이터 관리 — 엑셀 일괄 가져오기 / 전체 데이터 내보내기
"""
import streamlit as st
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).parent.parent))
from core.db import init_db, create_case, case_exists, next_case_number, get_all_cases
from core.ui_styles import inject_css, page_header

st.set_page_config(page_title="데이터 관리", page_icon="📥", layout="wide")
if "db_initialized" not in st.session_state:
    init_db()
    st.session_state["db_initialized"] = True

inject_css()
page_header("📥", "데이터 관리", "엑셀 가져오기 · 전체 데이터 내보내기")

tab_import, tab_export = st.tabs(["📥 가져오기 (엑셀 → DB)", "📤 내보내기 (DB → 엑셀)"])

# ══════════════════════════════════════════════
# 탭1: 가져오기
# ══════════════════════════════════════════════
with tab_import:
    st.markdown("#### 1단계 — 양식 다운로드")
    col_tmpl, _ = st.columns([2, 3])
    with col_tmpl:
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            import io

            def _make_template() -> bytes:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "사건목록"
                headers = [
                    "접수번호", "접수연도", "지역", "건물명", "건물소재지",
                    "신청인_성명", "신청인_주소", "신청인_연락처",
                    "피신청인_성명", "피신청인_주소", "피신청인_연락처", "피신청인_우편번호",
                    "신청내용", "분쟁유형", "접수일자", "조정동의여부",
                ]
                fill = PatternFill("solid", fgColor="1A56A0")
                font = Font(bold=True, color="FFFFFF")
                for col, h in enumerate(headers, 1):
                    c = ws.cell(1, col, h)
                    c.fill = fill
                    c.font = font
                    c.alignment = Alignment(horizontal="center")
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(len(h)+4, 14)
                ws.append(["2026-999", 2026, "수원시", "예시아파트", "경기도 수원시...",
                            "홍길동", "경기도 수원시...", "010-0000-0000",
                            "김철수", "경기도 성남시...", "010-1111-1111", "12345",
                            "관리비 미납", "관리비", "2026-01-01", "동의"])
                buf = io.BytesIO()
                wb.save(buf)
                return buf.getvalue()

            st.download_button(
                "📥 양식 다운로드 (.xlsx)",
                data=_make_template(),
                file_name="사건등록_양식.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except ImportError:
            st.warning("openpyxl이 설치되지 않아 양식 다운로드를 사용할 수 없습니다.")

    st.markdown("---")
    st.markdown("#### 2단계 — 파일 업로드 및 미리보기")

    uploaded = st.file_uploader("엑셀 파일 선택 (.xlsx)", type=["xlsx"],
                                 label_visibility="collapsed")

    if uploaded:
        try:
            import pandas as pd
            df = pd.read_excel(uploaded, dtype=str)
            df = df.fillna("")
            st.markdown(f"**{len(df)}행 감지됨**")
            st.dataframe(df.head(10), use_container_width=True, hide_index=True)

            col_dup, col_go = st.columns([2, 1])
            with col_dup:
                skip_dup = st.checkbox("이미 등록된 접수번호 건너뛰기", value=True)

            with col_go:
                if st.button("✅ 가져오기 실행", type="primary", use_container_width=True):
                    ok = err = dup = 0
                    log_lines = []
                    for _, row in df.iterrows():
                        case_no = str(row.get("접수번호", "")).strip()
                        if not case_no:
                            err += 1
                            log_lines.append("❌ 접수번호 없음 (행 건너뜀)")
                            continue
                        if case_exists(case_no):
                            if skip_dup:
                                dup += 1
                                log_lines.append(f"⏭ {case_no} — 이미 존재 (건너뜀)")
                                continue
                        try:
                            data = {k: (row.get(k, "") or "") for k in [
                                "접수번호", "접수연도", "지역", "건물명", "건물소재지",
                                "신청인_성명", "신청인_주소", "신청인_연락처",
                                "피신청인_성명", "피신청인_주소", "피신청인_연락처", "피신청인_우편번호",
                                "신청내용", "분쟁유형", "접수일자", "조정동의여부",
                            ]}
                            if not data.get("접수연도"):
                                import re
                                m = re.match(r"(\d{4})", case_no)
                                data["접수연도"] = int(m.group(1)) if m else 2026
                            else:
                                try:
                                    data["접수연도"] = int(data["접수연도"])
                                except Exception:
                                    data["접수연도"] = 2026
                            create_case(data)
                            ok += 1
                            log_lines.append(f"✅ {case_no} — 등록 완료")
                        except Exception as e:
                            err += 1
                            log_lines.append(f"❌ {case_no} — 오류: {e}")

                    st.success(f"완료: 등록 {ok}건 | 중복 {dup}건 | 오류 {err}건")
                    with st.expander("상세 로그"):
                        st.text("\n".join(log_lines))

        except Exception as e:
            st.error(f"파일 읽기 오류: {e}")

    st.markdown("---")
    st.markdown("""
**컬럼 안내**

| 컬럼명 | 필수 | 설명 |
|--------|------|------|
| 접수번호 | ✓ | 예: 2026-001 |
| 접수연도 | | 비어 있으면 접수번호에서 자동 추출 |
| 지역 / 건물명 / 건물소재지 | | 건물 기본 정보 |
| 신청인_성명·주소·연락처 | ✓ | |
| 피신청인_성명·주소·연락처·우편번호 | ✓ | |
| 신청내용 / 분쟁유형 | | |
| 접수일자 | | YYYY-MM-DD 형식 |
| 조정동의여부 | | 동의 / 미동의 |
""")


# ══════════════════════════════════════════════
# 탭2: 내보내기
# ══════════════════════════════════════════════
with tab_export:
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from datetime import date

    st.markdown("#### 내보내기 옵션")

    ec1, ec2, ec3 = st.columns(3)
    with ec1:
        year_now = date.today().year
        export_year = st.selectbox(
            "접수연도",
            ["전체"] + [str(y) for y in range(year_now, year_now - 6, -1)],
        )
    with ec2:
        export_status = st.selectbox(
            "진행상태",
            ["전체", "진행 중", "종결"],
        )
    with ec3:
        export_cols = st.multiselect(
            "포함 컬럼",
            ["기본정보", "신청인정보", "피신청인정보", "건물정보", "진행정보"],
            default=["기본정보", "신청인정보", "피신청인정보", "건물정보", "진행정보"],
        )

    if st.button("📤 엑셀 내보내기", type="primary", use_container_width=False):
        year_filter = None if export_year == "전체" else int(export_year)
        all_rows = [dict(r) for r in get_all_cases(year=year_filter)]

        # 진행상태 필터
        if export_status == "진행 중":
            all_rows = [r for r in all_rows if r.get("결과") not in ("조정성립", "조정불성립", "조정중지", "종결")]
        elif export_status == "종결":
            all_rows = [r for r in all_rows if r.get("결과") in ("조정성립", "조정불성립", "조정중지", "종결")]

        if not all_rows:
            st.warning("내보낼 데이터가 없습니다.")
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "사건목록"

            # 컬럼 정의
            COLS = []
            if "기본정보" in export_cols:
                COLS += [
                    ("접수번호",    "접수번호"),
                    ("접수연도",    "접수연도"),
                    ("접수일자",    "접수일자"),
                    ("지역",        "지역"),
                    ("분쟁유형",    "분쟁유형"),
                ]
            if "신청인정보" in export_cols:
                COLS += [
                    ("신청인_성명",     "신청인 성명"),
                    ("신청인_지위",     "신청인 지위"),
                    ("신청인_주소",     "신청인 주소"),
                    ("신청인_우편번호", "신청인 우편번호"),
                    ("신청인_연락처",   "신청인 연락처"),
                ]
            if "피신청인정보" in export_cols:
                COLS += [
                    ("피신청인_성명",     "피신청인 성명"),
                    ("피신청인_지위",     "피신청인 지위"),
                    ("피신청인_주소",     "피신청인 주소"),
                    ("피신청인_우편번호", "피신청인 우편번호"),
                    ("피신청인_연락처",   "피신청인 연락처"),
                ]
            if "건물정보" in export_cols:
                COLS += [
                    ("건물명",       "건물명"),
                    ("건물소재지",   "건물소재지"),
                    ("건축물용도",   "건축물용도"),
                ]
            if "진행정보" in export_cols:
                COLS += [
                    ("안내도달일",    "안내도달일"),
                    ("회신기한",      "회신기한"),
                    ("회신접수일",    "회신접수일"),
                    ("조정동의여부",  "조정동의여부"),
                    ("개최여부",      "개최여부"),
                    ("결과",          "결과"),
                    ("종결일자",      "종결일자"),
                ]

            # 헤더
            BLUE_FILL = PatternFill("solid", fgColor="1A56A0")
            HDR_FONT  = Font(bold=True, color="FFFFFF")
            thin      = Side(style="thin")
            BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)
            CENTER    = Alignment(horizontal="center", vertical="center")

            for c_idx, (_, label) in enumerate(COLS, 1):
                cell = ws.cell(row=1, column=c_idx, value=label)
                cell.fill      = BLUE_FILL
                cell.font      = HDR_FONT
                cell.alignment = CENTER
                cell.border    = BORDER

            # 데이터 행
            for r_idx, row in enumerate(all_rows, 2):
                for c_idx, (field, _) in enumerate(COLS, 1):
                    val  = row.get(field, "") or ""
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.alignment = Alignment(vertical="center")
                    cell.border    = BORDER

            # 컬럼 너비 자동
            for c_idx, (_, label) in enumerate(COLS, 1):
                col_letter = openpyxl.utils.get_column_letter(c_idx)
                ws.column_dimensions[col_letter].width = max(len(label) + 4, 12)

            ws.freeze_panes = "A2"

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            fname = f"사건데이터_{export_year}_{date.today()}.xlsx"
            st.success(f"✅ {len(all_rows)}건 내보내기 준비 완료")
            st.download_button(
                "⬇️ 다운로드",
                data=buf.read(),
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
