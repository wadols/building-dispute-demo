"""
신규 사건 접수 페이지
- Daum 주소 API (버튼 → 인라인 검색창 → 자동 입력)
- 신청인 지위: 텍스트 자유 입력
- 저장 후 st.rerun() → 접수번호 자동 갱신
"""
import streamlit as st
from datetime import date, timedelta
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).parent.parent))
from core.db import init_db, create_case, case_exists, next_case_number, get_case, update_case
from core.status_resolver import resolve_status
from core.ui_styles import inject_css, page_header, status_badge, section_header, case_folder_path
from core.address_search import address_search_widget

st.set_page_config(page_title="신규 접수", page_icon="📋", layout="wide")
if "db_initialized" not in st.session_state:
    init_db()
    st.session_state["db_initialized"] = True

inject_css()
OUTPUT_ROOT = Path(__file__).parent.parent / "output" / "사건자료"

# ── 모드 선택
# 외부(접수대장·사건상세)에서 edit_case를 주입하면 _active_edit에 보관 (reruns 에도 유지)
if "edit_case" in st.session_state:
    st.session_state["_active_edit"] = st.session_state.pop("edit_case")

_active_edit = st.session_state.get("_active_edit")
mode = "기존 사건 수정" if _active_edit else "신규 접수"

existing = None
if mode == "기존 사건 수정":
    if _active_edit:
        edit_num = _active_edit
    else:
        edit_num = st.sidebar.text_input("수정할 접수번호", key="_edit_num_input")
    if edit_num:
        row = get_case(edit_num.strip())
        if row is None:
            st.sidebar.error("존재하지 않는 접수번호입니다.")
            st.stop()
        existing = dict(row)
    else:
        st.sidebar.info("접수번호를 입력하세요.")
        st.stop()

today = date.today()

# ────────────────────────────────────────────────
# session_state 초기화
# ────────────────────────────────────────────────
# 신규 모드: 매번 접수번호를 최신으로 계산
# 수정 모드: 해당 사건 값으로 1회 초기화
_init_key = f"_form_ready_{'new' if not existing else existing['접수번호']}"

if _init_key not in st.session_state:
    def _v(f, default=""):
        return existing[f] if (existing and existing.get(f) is not None) else default

    def _d(val: str) -> date | None:
        if not val:
            return None
        return date.fromisoformat(str(val)[:10])

    st.session_state["inp_접수번호"]          = _v("접수번호", next_case_number(today.year))
    st.session_state["inp_접수일자"]          = _d(_v("접수일자", today.isoformat()))
    st.session_state["inp_지역"]              = _v("지역")
    st.session_state["inp_분쟁유형"]          = _v("분쟁유형")
    st.session_state["inp_신청내용"]          = _v("신청내용")
    st.session_state["inp_건물명"]            = _v("건물명")
    st.session_state["inp_건물소재지"]        = _v("건물소재지")
    st.session_state["inp_건축물용도"]        = _v("건축물용도")
    st.session_state["inp_신청인_성명"]       = _v("신청인_성명")
    st.session_state["inp_신청인_주소"]       = _v("신청인_주소")
    st.session_state["inp_신청인_우편번호"]   = _v("신청인_우편번호")
    st.session_state["inp_신청인_연락처"]     = _v("신청인_연락처")
    st.session_state["inp_신청인_지위"]       = _v("신청인_지위")
    st.session_state["inp_피신청인_성명"]     = _v("피신청인_성명")
    st.session_state["inp_피신청인_주소"]     = _v("피신청인_주소")
    st.session_state["inp_피신청인_우편번호"] = _v("피신청인_우편번호")
    st.session_state["inp_피신청인_연락처"]   = _v("피신청인_연락처")
    st.session_state["inp_피신청인_지위"]     = _v("피신청인_지위")
    st.session_state["inp_피신청인2_성명"]     = _v("피신청인2_성명")
    st.session_state["inp_피신청인2_주소"]     = _v("피신청인2_주소")
    st.session_state["inp_피신청인2_우편번호"] = _v("피신청인2_우편번호")
    st.session_state["inp_피신청인2_연락처"]   = _v("피신청인2_연락처")
    st.session_state["inp_피신청인2_지위"]     = _v("피신청인2_지위")
    st.session_state.setdefault("show_rp2", bool(existing and existing.get("피신청인2_성명")))
    st.session_state.setdefault("show_rp2_search", False)
    st.session_state["inp_안내도달일"]        = _d(_v("안내도달일"))
    st.session_state["inp_회신기한"]          = _d(_v("회신기한"))
    st.session_state["inp_회신접수일"]        = _d(_v("회신접수일"))
    st.session_state["inp_조정동의여부"]      = _v("조정동의여부")
    st.session_state["inp_개최여부"]          = _v("개최여부")
    st.session_state["inp_결과"]              = _v("결과")
    st.session_state["inp_종결일자"]          = _d(_v("종결일자"))
    # 주소 검색 표시 여부
    st.session_state.setdefault("show_ap_search", False)
    st.session_state.setdefault("show_rp_search", False)
    st.session_state[_init_key] = True

# ────────────────────────────────────────────────
# 주소 검색 결과 적용 함수
# ────────────────────────────────────────────────
def apply_addr(result, addr_key: str, zip_key: str, show_key: str):
    """컴포넌트 반환값 → 주소·우편번호 갱신 + 창 자동 닫기"""
    prev_k = f"_prev_{show_key}"
    if result and result != st.session_state.get(prev_k):
        st.session_state[prev_k]   = result
        st.session_state[addr_key] = result["address"]
        st.session_state[zip_key]  = result["postcode"]
        st.session_state[show_key] = False   # ← 정확한 키로 창 닫기
        st.rerun()

# ── 헤더
if mode == "신규 접수":
    page_header("📋", "신규 사건 접수", f"접수연도: {today.year}")
else:
    page_header("✏️", f"사건 수정 — {existing['접수번호']}", existing.get("신청인_성명",""))

# ── 탭
tab_single, tab_bulk = st.tabs(["📋 단건 접수", "📥 엑셀 일괄 가져오기"])

# ══════════════════════════════════════════════════════════════
# 탭 1: 단건 접수 (기존 폼)
# ══════════════════════════════════════════════════════════════
with tab_single:

    # ── PDF 자동 입력 섹션
    with st.expander("📄 신청서 PDF에서 자동 입력", expanded=False):

        # rerun 후에도 결과 메시지 유지
        _pdf_msg = st.session_state.pop("_pdf_msg", None)
        if _pdf_msg:
            if _pdf_msg.get("filled"):
                st.success(f"✅ 자동 입력 완료: {', '.join(_pdf_msg['filled'])}")
            if _pdf_msg.get("skipped"):
                st.info(f"직접 입력 필요: {', '.join(_pdf_msg['skipped'])}")
            if _pdf_msg.get("error"):
                st.error(_pdf_msg["error"])
            if _pdf_msg.get("raw"):
                with st.expander("추출 원문 확인 (파싱이 안 될 때 참고)"):
                    st.text(_pdf_msg["raw"][:3000])

        pdf_file = st.file_uploader(
            "신청서 PDF 업로드 (집합건물 분쟁조정 신청서)",
            type=["pdf"], key="pdf_upload",
            label_visibility="collapsed",
        )
        if pdf_file:
            if st.button("🔍 자동 추출 시작", key="pdf_parse_btn", type="primary"):
                msg = {}
                try:
                    from core.pdf_parser import parse_application_pdf, FIELD_LABELS
                    parsed, raw_text = parse_application_pdf(pdf_file.read())
                    msg["raw"] = raw_text

                    if "_error" in parsed:
                        msg["error"] = f"파싱 오류: {parsed['_error']}"
                    else:
                        field_map = {
                            "신청인_성명":    "inp_신청인_성명",
                            "신청인_주소":    "inp_신청인_주소",
                            "신청인_연락처":  "inp_신청인_연락처",
                            "피신청인_성명":  "inp_피신청인_성명",
                            "피신청인_주소":  "inp_피신청인_주소",
                            "피신청인_연락처":"inp_피신청인_연락처",
                            "지역":           "inp_지역",
                            "건물소재지":     "inp_건물소재지",
                        }
                        filled, skipped = [], []
                        for src, dst in field_map.items():
                            if parsed.get(src):
                                st.session_state[dst] = parsed[src]
                                filled.append(FIELD_LABELS.get(src, src))
                            else:
                                skipped.append(FIELD_LABELS.get(src, src))

                        if parsed.get("접수일자"):
                            try:
                                st.session_state["inp_접수일자"] = date.fromisoformat(parsed["접수일자"])
                                filled.append("접수일자")
                            except Exception:
                                pass

                        # 우편번호 자동 조회 (카카오 로컬 API)
                        from core.kakao_api import lookup_postcode
                        for addr_key, zip_key, label in [
                            ("신청인_주소",  "inp_신청인_우편번호",  "신청인 우편번호"),
                            ("피신청인_주소","inp_피신청인_우편번호","피신청인 우편번호"),
                        ]:
                            if parsed.get(addr_key):
                                postcode = lookup_postcode(parsed[addr_key])
                                if postcode:
                                    st.session_state[zip_key] = postcode
                                    filled.append(label)

                        if not raw_text.strip():
                            msg["error"] = "PDF에서 텍스트를 읽지 못했습니다. 스캔 이미지 PDF이거나 보안이 걸린 파일일 수 있습니다."
                        elif not filled:
                            msg["error"] = "텍스트는 읽었으나 항목을 추출하지 못했습니다. '추출 원문 확인'으로 내용을 확인하세요."
                        msg["filled"] = filled
                        msg["skipped"] = skipped

                except ImportError:
                    msg["error"] = "pdfplumber가 필요합니다. 터미널에서: pip install pdfplumber"
                except Exception as e:
                    msg["error"] = f"오류: {e}"

                st.session_state["_pdf_msg"] = msg
                st.rerun()
        else:
            st.caption("신청인이 제출한 '집합건물 분쟁조정 신청서' PDF를 업로드하면 성명·주소·연락처·신청내용 등을 자동으로 입력합니다.")

    st.markdown("")

    # 신규 접수: 접수번호가 비어 있으면 항상 최신 번호로 채움
    if mode == "신규 접수" and not st.session_state.get("inp_접수번호"):
        st.session_state["inp_접수번호"] = next_case_number(today.year)

    # ══════════════════════════════════════════════
    # ① 사건 기본 정보
    # ══════════════════════════════════════════════
    with st.container(border=True):
        section_header("①", "사건 기본 정보")
        c1, c2, c3 = st.columns([2, 2, 2])
        with c1:
            st.text_input("접수번호 *", key="inp_접수번호", help="자동 생성 — 필요시 수정 가능")
        with c2:
            st.date_input("접수일자 *", key="inp_접수일자")
        with c3:
            st.text_input("지역 *", key="inp_지역", placeholder="예: 수원시 영통구")
        분쟁유형_opts = [
            "",
            "하자",
            "관리인·관리위원 선임·해임 및 관리단 구성·운영",
            "공용부분 보존·관리·변경",
            "관리비 징수·관리·사용",
            "규약 제정·개정",
            "재건축 관련 철거·비용분담·구분소유권 귀속",
            "소음·약취 등 공동생활",
            "대지·부속시설 보존·관리·변경",
            "전유부분 사용방법",
            "관리비 외 수입 징수·관리·사용",
            "관리위탁계약 등 관리단 체결 계약",
            "기타",
        ]
        c4, c5 = st.columns([2, 4])
        with c4:
            st.selectbox("분쟁유형", 분쟁유형_opts, key="inp_분쟁유형")
        with c5:
            st.text_area("신청내용", key="inp_신청내용", height=72)

    with st.container(border=True):
        section_header("②", "건물 정보")
        b1, b2, b3 = st.columns([2, 3, 2])
        용도_opts = ["", "아파트", "연립주택", "다세대주택", "오피스텔", "상가", "복합건물", "기타"]
        with b1:
            st.text_input("건물명", key="inp_건물명", placeholder="예: 행복아파트")
        with b2:
            st.text_input("건물소재지", key="inp_건물소재지")
        with b3:
            st.selectbox("건축물용도", 용도_opts, key="inp_건축물용도")

    col_ap, col_rp = st.columns(2)

    with col_ap:
        with st.container(border=True):
            section_header("③", "신청인 정보")
            st.text_input("성명 *", key="inp_신청인_성명")
            st.text_input("지위", key="inp_신청인_지위", placeholder="예: 구분소유자, 임차인 등 자유 입력")
            st.text_input("연락처", key="inp_신청인_연락처", placeholder="010-0000-0000")
            if st.button("🔍 주소 검색 (다음)", key="btn_ap_search", use_container_width=True):
                st.session_state["show_ap_search"] = not st.session_state.get("show_ap_search", False)
            if st.session_state.get("show_ap_search", False):
                ap_result = address_search_widget(key="ap_addr_search")
                apply_addr(ap_result,
                           addr_key="inp_신청인_주소", zip_key="inp_신청인_우편번호",
                           show_key="show_ap_search")
            st.text_input("주소 *", key="inp_신청인_주소")
            st.text_input("우편번호", key="inp_신청인_우편번호", max_chars=6)

    with col_rp:
        with st.container(border=True):
            section_header("④", "피신청인 정보")
            st.text_input("성명 *", key="inp_피신청인_성명")
            st.text_input("지위", key="inp_피신청인_지위", placeholder="예: 관리단, 입주자대표회의 등 자유 입력")
            st.text_input("연락처", key="inp_피신청인_연락처", placeholder="010-0000-0000")
            if st.button("🔍 주소 검색 (다음)", key="btn_rp_search", use_container_width=True):
                st.session_state["show_rp_search"] = not st.session_state.get("show_rp_search", False)
            if st.session_state.get("show_rp_search", False):
                rp_result = address_search_widget(key="rp_addr_search")
                apply_addr(rp_result,
                           addr_key="inp_피신청인_주소", zip_key="inp_피신청인_우편번호",
                           show_key="show_rp_search")
            st.text_input("주소 *", key="inp_피신청인_주소")
            st.text_input("우편번호", key="inp_피신청인_우편번호", max_chars=6)
            if not st.session_state.get("show_rp2"):
                st.markdown('<div style="height:4px"></div>', unsafe_allow_html=True)
                if st.button("➕ 피신청인 추가", use_container_width=True, key="btn_add_rp2"):
                    st.session_state["show_rp2"] = True
                    st.rerun()

    # ── 피신청인2 (조건부 표시)
    if st.session_state.get("show_rp2"):
        with st.container(border=True):
            hdr_col, del_col = st.columns([9, 1])
            with hdr_col:
                section_header("④-2", "피신청인2 정보")
            with del_col:
                st.markdown('<div style="margin-top:6px"></div>', unsafe_allow_html=True)
                if st.button("✕ 제거", key="btn_del_rp2", use_container_width=True):
                    st.session_state["show_rp2"] = False
                    for _k in ["inp_피신청인2_성명", "inp_피신청인2_주소",
                               "inp_피신청인2_우편번호", "inp_피신청인2_연락처", "inp_피신청인2_지위"]:
                        st.session_state[_k] = ""
                    st.session_state["show_rp2_search"] = False
                    st.rerun()

            rp2a, rp2b = st.columns(2)
            with rp2a:
                st.text_input("성명 *", key="inp_피신청인2_성명", placeholder="두번째 피신청인")
                st.text_input("지위", key="inp_피신청인2_지위", placeholder="예: 관리단, 입주자대표회의")
                st.text_input("연락처", key="inp_피신청인2_연락처", placeholder="010-0000-0000")
            with rp2b:
                if st.button("🔍 주소 검색 (다음)", key="btn_rp2_search", use_container_width=True):
                    st.session_state["show_rp2_search"] = not st.session_state.get("show_rp2_search", False)
                if st.session_state.get("show_rp2_search", False):
                    rp2_result = address_search_widget(key="rp2_addr_search")
                    apply_addr(rp2_result,
                               addr_key="inp_피신청인2_주소", zip_key="inp_피신청인2_우편번호",
                               show_key="show_rp2_search")
                st.text_input("주소", key="inp_피신청인2_주소")
                st.text_input("우편번호", key="inp_피신청인2_우편번호", max_chars=6)

    # ══════════════════════════════════════════════
    # ⑤ 진행 정보
    # ══════════════════════════════════════════════
    with st.container(border=True):
        section_header("⑤", "진행 정보")
        d1, d2, d3 = st.columns([2, 1, 2])
        with d1:
            st.date_input("안내도달일", value=None, key="inp_안내도달일")
        with d2:
            deadline_days = st.number_input("회신기한 (일)", min_value=1, max_value=60, value=7)
        with d3:
            안내도달일_val = st.session_state.get("inp_안내도달일")
            if isinstance(안내도달일_val, date):
                auto_dl = 안내도달일_val + timedelta(days=int(deadline_days))
                if st.session_state.get("inp_회신기한") != auto_dl:
                    st.session_state["inp_회신기한"] = auto_dl
            st.date_input("회신기한 ✦자동", key="inp_회신기한")

        d4, d5, d6, d7 = st.columns(4)
        with d4:
            st.date_input("회신접수일", value=None, key="inp_회신접수일")
        with d5:
            st.selectbox("조정동의여부", ["", "동의", "부동의", "무응답"], key="inp_조정동의여부")
        with d6:
            st.selectbox("개최여부", ["", "개최", "불개시"], key="inp_개최여부")
        with d7:
            st.selectbox("결과", ["", "조정성립", "조정불성립", "조정중지", "종결"], key="inp_결과")

        st.date_input("종결일자", value=None, key="inp_종결일자")

    # ══════════════════════════════════════════════
    # 저장 버튼
    # ══════════════════════════════════════════════
    st.markdown('<div style="height:6px"></div>', unsafe_allow_html=True)
    submitted = st.button(
        "✅  저장 완료" if mode == "신규 접수" else "💾  수정 저장",
        use_container_width=True,
        type="primary",
    )

    if submitted:
        def gs(k):      return st.session_state.get(k)
        def gss(k):
            v = gs(k)
            return v.strip() if isinstance(v, str) else (str(v).strip() if v else "")
        def gsd(k):
            v = gs(k)
            return v.isoformat() if isinstance(v, date) else None
        def opt(v):     return v or None

        # 필수 검증
        errs = []
        if not gss("inp_접수번호"):     errs.append("접수번호를 입력하세요.")
        if not gss("inp_지역"):         errs.append("지역을 입력하세요.")
        if not gss("inp_신청인_성명"):  errs.append("신청인 성명을 입력하세요.")
        if not gss("inp_신청인_주소"):  errs.append("신청인 주소를 입력하세요.")
        if not gss("inp_피신청인_성명"):errs.append("피신청인 성명을 입력하세요.")
        if not gss("inp_피신청인_주소"):errs.append("피신청인 주소를 입력하세요.")
        if errs:
            for e in errs:
                st.error(e)
            st.stop()

        번호 = gss("inp_접수번호")
        접수일자_v = gs("inp_접수일자")
        if not isinstance(접수일자_v, date):
            st.error("접수일자를 선택하세요.")
            st.stop()

        if mode == "신규 접수" and case_exists(번호):
            st.error(f"**{번호}** 는 이미 존재합니다. '기존 사건 수정' 모드를 사용하세요.")
            st.stop()

        # 접수연도는 접수번호 앞자리 연도 기준 (접수일자가 전년 12월일 수 있음)
        try:
            _yr = int(번호.split("-")[0])
        except Exception:
            _yr = 접수일자_v.year

        data = {
            "접수연도":          _yr,
            "접수번호":          번호,
            "지역":              gss("inp_지역"),
            "신청인_성명":       gss("inp_신청인_성명"),
            "신청인_주소":       gss("inp_신청인_주소"),
            "신청인_우편번호":   opt(gss("inp_신청인_우편번호")),
            "신청인_연락처":     opt(gss("inp_신청인_연락처")),
            "신청인_지위":       opt(gss("inp_신청인_지위")),
            "피신청인_성명":     gss("inp_피신청인_성명"),
            "피신청인_주소":     gss("inp_피신청인_주소"),
            "피신청인_우편번호": opt(gss("inp_피신청인_우편번호")),
            "피신청인_연락처":   opt(gss("inp_피신청인_연락처")),
            "피신청인_지위":     opt(gss("inp_피신청인_지위")),
            "피신청인2_성명":     opt(gss("inp_피신청인2_성명")) if st.session_state.get("show_rp2") else None,
            "피신청인2_주소":     opt(gss("inp_피신청인2_주소")) if st.session_state.get("show_rp2") else None,
            "피신청인2_우편번호": opt(gss("inp_피신청인2_우편번호")) if st.session_state.get("show_rp2") else None,
            "피신청인2_연락처":   opt(gss("inp_피신청인2_연락처")) if st.session_state.get("show_rp2") else None,
            "피신청인2_지위":     opt(gss("inp_피신청인2_지위")) if st.session_state.get("show_rp2") else None,
            "건물명":            opt(gss("inp_건물명")),
            "건물소재지":        opt(gss("inp_건물소재지")),
            "건축물용도":        opt(gs("inp_건축물용도") or None),
            "신청내용":          opt(gss("inp_신청내용")),
            "접수일자":          접수일자_v.isoformat(),
            "안내도달일":        gsd("inp_안내도달일"),
            "회신기한":          gsd("inp_회신기한"),
            "회신접수일":        gsd("inp_회신접수일"),
            "조정동의여부":      opt(gs("inp_조정동의여부") or None),
            "개최여부":          opt(gs("inp_개최여부") or None),
            "결과":              opt(gs("inp_결과") or None),
            "종결일자":          gsd("inp_종결일자"),
            "분쟁유형":          opt(gs("inp_분쟁유형") or None),
        }
        data["진행상태"] = resolve_status(data)

        try:
            if mode == "신규 접수":
                create_case(data)
                folder = case_folder_path(
                    OUTPUT_ROOT, 번호,
                    data.get("신청인_성명", ""), data.get("피신청인_성명", ""),
                )
                folder.mkdir(parents=True, exist_ok=True)
                # 세션 정리 후 접수대장으로 이동
                for k in [k for k in list(st.session_state) if k.startswith("inp_") or k == _init_key]:
                    del st.session_state[k]
                st.cache_data.clear()
                st.toast(f"✅ {번호} 접수 완료!", icon="✅")
                st.switch_page("pages/3_접수대장.py")
            else:
                data.pop("접수번호", None)
                update_case(번호, data)
                st.session_state.pop("_active_edit", None)
                origin = st.session_state.pop("_edit_origin", None)
                origin_case = st.session_state.pop("_edit_origin_case", None)
                if origin == "pages/4_사건상세.py" and origin_case:
                    st.session_state["detail_case"] = origin_case
                    st.cache_data.clear()
                    st.switch_page("pages/4_사건상세.py")
                else:
                    st.cache_data.clear()
                    st.switch_page("pages/3_접수대장.py")
        except Exception as e:
            st.error(f"저장 오류: {e}")

    st.markdown(
        '<p style="font-size:0.8rem;color:#94A3B8;margin-top:12px;">'
        '✦ 주소 검색 버튼 클릭 → 다음 주소창 → 선택 시 자동 입력됩니다. &nbsp;'
        '✦ 회신기한은 안내도달일 + 설정 일수로 자동 계산됩니다.'
        '</p>',
        unsafe_allow_html=True,
    )

# ══════════════════════════════════════════════════════════════
# 탭 2: 엑셀 일괄 가져오기
# ══════════════════════════════════════════════════════════════
with tab_bulk:
    st.markdown("#### 1단계 — 양식 다운로드")
    col_tmpl, _ = st.columns([2, 3])
    with col_tmpl:
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            import io

            def _make_template() -> bytes:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "사건목록"
                headers = [
                    "접수번호", "접수연도", "지역", "건물명", "건물소재지",
                    "신청인_성명", "신청인_주소", "신청인_연락처",
                    "피신청인_성명", "피신청인_주소", "피신청인_연락처", "피신청인_우편번호",
                    "신청내용", "분쟁유형", "접수일자", "조정동의여부",
                ]
                fill = PatternFill("solid", fgColor="1A56A0")
                font = Font(bold=True, color="FFFFFF")
                for col, h in enumerate(headers, 1):
                    c = ws.cell(1, col, h)
                    c.fill = fill
                    c.font = font
                    c.alignment = Alignment(horizontal="center")
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max(len(h)+4, 14)
                ws.append(["2026-999", 2026, "수원시", "예시아파트", "경기도 수원시...",
                            "홍길동", "경기도 수원시...", "010-0000-0000",
                            "김철수", "경기도 성남시...", "010-1111-1111", "12345",
                            "관리비 미납", "관리비", "2026-01-01", "동의"])
                buf = io.BytesIO()
                wb.save(buf)
                return buf.getvalue()

            st.download_button(
                "📥 양식 다운로드 (.xlsx)",
                data=_make_template(),
                file_name="사건등록_양식.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except ImportError:
            st.warning("openpyxl이 설치되지 않아 양식 다운로드를 사용할 수 없습니다.")

    st.markdown("---")
    st.markdown("#### 2단계 — 파일 업로드 및 미리보기")

    uploaded = st.file_uploader("엑셀 파일 선택 (.xlsx)", type=["xlsx"],
                                label_visibility="collapsed", key="bulk_upload")

    if uploaded:
        try:
            import pandas as pd
            df = pd.read_excel(uploaded, dtype=str)
            df = df.fillna("")
            st.markdown(f"**{len(df)}행 감지됨**")
            st.dataframe(df.head(10), use_container_width=True, hide_index=True)

            col_dup, col_go = st.columns([2, 1])
            with col_dup:
                skip_dup = st.checkbox("이미 등록된 접수번호 건너뛰기", value=True, key="bulk_skip_dup")

            with col_go:
                if st.button("✅ 가져오기 실행", type="primary", use_container_width=True, key="bulk_run"):
                    import re as _re

                    def _cell(v):
                        s = str(v).strip() if v is not None else ""
                        return s  # NOT NULL 컬럼은 빈 문자열로 허용

                    # CHECK 제약이 있는 선택 컬럼: 빈 값 → None
                    _nullable_check = {"조정동의여부"}

                    ok = err = dup = 0
                    log_lines = []
                    for _, row in df.iterrows():
                        case_no = str(row.get("접수번호", "")).strip()
                        if not case_no:
                            err += 1
                            log_lines.append("❌ 접수번호 없음 (행 건너뜀)")
                            continue
                        if case_exists(case_no):
                            if skip_dup:
                                dup += 1
                                log_lines.append(f"⏭ {case_no} — 이미 존재 (건너뜀)")
                                continue
                        try:
                            data = {k: _cell(row.get(k)) for k in [
                                "접수번호", "접수연도", "지역", "건물명", "건물소재지",
                                "신청인_성명", "신청인_주소", "신청인_연락처",
                                "피신청인_성명", "피신청인_주소", "피신청인_연락처", "피신청인_우편번호",
                                "신청내용", "분쟁유형", "접수일자", "조정동의여부",
                            ]}
                            for _col in _nullable_check:
                                if not data.get(_col):
                                    data[_col] = None
                            if not data.get("접수연도"):
                                m = _re.match(r"(\d{4})", case_no)
                                data["접수연도"] = int(m.group(1)) if m else today.year
                            else:
                                try:
                                    data["접수연도"] = int(data["접수연도"])
                                except Exception:
                                    data["접수연도"] = today.year
                            create_case(data)
                            ok += 1
                            log_lines.append(f"✅ {case_no} — 등록 완료")
                        except Exception as e:
                            err += 1
                            log_lines.append(f"❌ {case_no} — 오류: {e}")

                    st.success(f"완료: 등록 {ok}건 | 중복 {dup}건 | 오류 {err}건")
                    with st.expander("상세 로그"):
                        st.text("\n".join(log_lines))

        except Exception as e:
            st.error(f"파일 읽기 오류: {e}")

    st.markdown("---")
    st.markdown("""
**컬럼 안내**

| 컬럼명 | 필수 | 설명 |
|--------|------|------|
| 접수번호 | ✓ | 예: 2026-001 |
| 접수연도 | | 비어 있으면 접수번호에서 자동 추출 |
| 지역 / 건물명 / 건물소재지 | | 건물 기본 정보 |
| 신청인_성명·주소·연락처 | ✓ | |
| 피신청인_성명·주소·연락처·우편번호 | ✓ | |
| 신청내용 / 분쟁유형 | | |
| 접수일자 | | YYYY-MM-DD 형식 |
| 조정동의여부 | | 동의 / 미동의 |
""")
