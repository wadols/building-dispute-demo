"""
보고서 — 월별/분기별/연도별 통계 및 차트
"""
import streamlit as st
from datetime import date, timedelta
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).parent.parent))
from core.db import init_db, get_stats_by_period, get_monthly_counts, get_conn
from core.ui_styles import inject_css, page_header

st.set_page_config(page_title="보고서", page_icon="📊", layout="wide")
if "db_initialized" not in st.session_state:
    init_db()
    st.session_state["db_initialized"] = True

inject_css()
page_header("📊", "보고서", "월별·분기별·연도별 업무 통계")

# ── 기간 선택
col_type, col_year, col_q, col_m = st.columns([2, 1, 1, 1])
with col_type:
    period_type = st.radio("집계 단위", ["월별", "분기별", "연도별", "기간 직접설정"],
                           index=2, horizontal=True, label_visibility="collapsed")
today = date.today()
year_now = today.year

with col_year:
    sel_year = st.selectbox("연도", list(range(year_now, year_now - 5, -1)), index=0,
                             label_visibility="collapsed")
with col_q:
    sel_q = st.selectbox("분기", ["1분기", "2분기", "3분기", "4분기"],
                         index=(today.month - 1) // 3, label_visibility="collapsed")
with col_m:
    sel_m = st.selectbox("월", [f"{i}월" for i in range(1, 13)],
                         index=today.month - 1, label_visibility="collapsed")

# 기간 계산
if period_type == "월별":
    m = int(sel_m.replace("월", ""))
    start_d = date(sel_year, m, 1)
    next_m = m % 12 + 1
    next_y = sel_year + (1 if m == 12 else 0)
    end_d = date(next_y, next_m, 1) - timedelta(days=1)
    period_label = f"{sel_year}년 {m}월"
elif period_type == "분기별":
    q = int(sel_q[0])
    start_m = (q - 1) * 3 + 1
    end_m   = q * 3
    start_d = date(sel_year, start_m, 1)
    end_d   = date(sel_year, end_m, [31,28,31,30,31,30,31,31,30,31,30,31][end_m-1])
    if sel_year % 4 == 0 and end_m == 2:
        end_d = date(sel_year, 2, 29)
    period_label = f"{sel_year}년 {sel_q}"
elif period_type == "연도별":
    start_d = date(sel_year, 1, 1)
    end_d   = date(sel_year, 12, 31)
    period_label = f"{sel_year}년 전체"
else:
    c1, c2 = st.columns(2)
    with c1:
        start_d = st.date_input("시작일", value=date(today.year, 1, 1))
    with c2:
        end_d = st.date_input("종료일", value=today)
    period_label = f"{start_d} ~ {end_d}"

st.markdown(f"#### {period_label} 통계")
st.markdown("---")

# ── 통계 집계
stats = get_stats_by_period(str(start_d), str(end_d))

# 연도별 모드: 접수건수는 접수연도 기준으로 재집계 (접수일자가 전년 12월인 케이스 포함)
if period_type == "연도별":
    with get_conn() as _conn:
        stats["접수건수"] = _conn.execute(
            "SELECT COUNT(*) FROM cases WHERE 접수연도 = ?", (sel_year,)
        ).fetchone()[0]

# ── KPI 카드
k1, k2, k3, k4 = st.columns(4)
kpis = [
    (k1, "접수건수",   stats["접수건수"],                "#1A56A0", ""),
    (k2, "처리중",     stats["처리중"],                  "#E67E22", "accent"),
    (k3, "동의율",     f"{stats['동의율']}%",            "#27AE60", "green"),
    (k4, "조정성립률", f"{stats['성립률']}%",            "#E74C3C", "red"),
]
for col, label, val, _, accent in kpis:
    with col:
        st.markdown(
            f'<div class="kpi-card {accent}">'
            f'<div class="kpi-value">{val}</div>'
            f'<div class="kpi-label">{label}</div>'
            f'</div>', unsafe_allow_html=True)

st.markdown("")

# ── 차트 (plotly 없으면 테이블로 대체)
col_left, col_right = st.columns(2)

# 분쟁유형별
with col_left:
    st.markdown("##### 분쟁유형별 건수")
    dtype_data = stats.get("분쟁유형별", [])
    if dtype_data:
        try:
            import plotly.express as px
            import pandas as pd
            df = pd.DataFrame(dtype_data)
            fig = px.pie(df, names="분쟁유형", values="cnt",
                         color_discrete_sequence=px.colors.qualitative.Set2)
            fig.update_layout(margin=dict(t=10, b=10, l=10, r=10), height=280)
            st.plotly_chart(fig, use_container_width=True)
        except ImportError:
            for row in dtype_data:
                st.write(f"- {row['분쟁유형'] or '미분류'}: {row['cnt']}건")
    else:
        st.info("데이터 없음")

# 지역별
with col_right:
    st.markdown("##### 지역별 건수")
    region_data = stats.get("지역별", [])
    if region_data:
        try:
            import plotly.express as px
            import pandas as pd
            df = pd.DataFrame(region_data)
            fig = px.bar(df, x="지역", y="cnt", text="cnt",
                         color_discrete_sequence=["#1A56A0"])
            fig.update_layout(margin=dict(t=10, b=10, l=10, r=10), height=280,
                              xaxis_title="", yaxis_title="건수")
            st.plotly_chart(fig, use_container_width=True)
        except ImportError:
            for row in region_data:
                st.write(f"- {row['지역'] or '미지정'}: {row['cnt']}건")
    else:
        st.info("데이터 없음")

# ── 월별 추이 (연도별 모드에서만)
if period_type == "연도별":
    st.markdown("---")
    st.markdown("##### 월별 접수 건수")
    monthly = get_monthly_counts(sel_year)
    if monthly:
        try:
            import plotly.graph_objects as go
            import pandas as pd
            df = pd.DataFrame(monthly)
            fig = go.Figure()
            fig.add_trace(go.Bar(x=df["월"], y=df["접수"], name="접수", marker_color="#1A56A0",
                                 text=df["접수"], textposition="outside"))
            fig.update_layout(barmode="group", margin=dict(t=10,b=10,l=10,r=10),
                              height=300, xaxis=dict(tickvals=list(range(1,13)),
                              ticktext=[f"{i}월" for i in range(1,13)]),
                              showlegend=False)
            st.plotly_chart(fig, use_container_width=True)
        except ImportError:
            import pandas as pd
            st.dataframe(pd.DataFrame(monthly), use_container_width=True)
    else:
        st.info("데이터 없음")

# ── 엑셀 내보내기
st.markdown("---")
if st.button("📥 엑셀로 내보내기", type="primary"):
    import io, openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ── 공통 스타일 ──────────────────────────────
    BLUE      = "1A56A0"
    BLUE_FONT = Font(bold=True, color="FFFFFF")
    HDR_FILL  = PatternFill("solid", fgColor=BLUE)
    SUB_FILL  = PatternFill("solid", fgColor="D9E1F2")
    BOLD      = Font(bold=True)
    CENTER    = Alignment(horizontal="center", vertical="center")
    def thin_border():
        s = Side(style="thin")
        return Border(left=s, right=s, top=s, bottom=s)

    def write_header(ws, row, cols, fill=None):
        fill = fill or HDR_FILL
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font  = BLUE_FONT if fill == HDR_FILL else BOLD
            cell.fill  = fill
            cell.alignment = CENTER
            cell.border = thin_border()

    def write_row(ws, row, vals):
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.alignment = Alignment(vertical="center")
            cell.border    = thin_border()

    # ════════════════════════════════════════════
    # 시트1: 통계 요약
    # ════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "통계요약"

    # 제목
    ws1.merge_cells("A1:F1")
    t = ws1["A1"]
    t.value     = f"집합건물 분쟁조정위원회 업무 통계 — {period_label}"
    t.font      = Font(bold=True, size=14)
    t.alignment = CENTER
    ws1.row_dimensions[1].height = 28

    # KPI
    r = 3
    ws1.merge_cells(f"A{r}:F{r}")
    ws1[f"A{r}"].value = "▶ 주요 지표"
    ws1[f"A{r}"].font  = BOLD
    r += 1
    write_header(ws1, r, ["접수건수", "처리중", "종결", "동의율", "성립건수", "조정성립률"])
    r += 1
    write_row(ws1, r, [
        stats["접수건수"], stats["처리중"], stats["종결"],
        f"{stats['동의율']}%", stats["성립건수"], f"{stats['성립률']}%",
    ])
    r += 2

    # 분쟁유형별
    ws1.merge_cells(f"A{r}:B{r}")
    ws1[f"A{r}"].value = "▶ 분쟁유형별"
    ws1[f"A{r}"].font  = BOLD
    r += 1
    write_header(ws1, r, ["분쟁유형", "건수"], fill=SUB_FILL)
    r += 1
    for item in stats.get("분쟁유형별", []):
        write_row(ws1, r, [item.get("분쟁유형") or "미분류", item.get("cnt", 0)])
        r += 1
    r += 1

    # 지역별
    ws1.merge_cells(f"A{r}:B{r}")
    ws1[f"A{r}"].value = "▶ 지역별"
    ws1[f"A{r}"].font  = BOLD
    r += 1
    write_header(ws1, r, ["지역", "건수"], fill=SUB_FILL)
    r += 1
    for item in stats.get("지역별", []):
        write_row(ws1, r, [item.get("지역") or "미지정", item.get("cnt", 0)])
        r += 1
    r += 1

    # 월별 추이 (연도별 모드)
    if period_type == "연도별":
        monthly = get_monthly_counts(sel_year)
        if monthly:
            ws1.merge_cells(f"A{r}:C{r}")
            ws1[f"A{r}"].value = "▶ 월별 접수·종결 추이"
            ws1[f"A{r}"].font  = BOLD
            r += 1
            write_header(ws1, r, ["월", "접수", "종결"], fill=SUB_FILL)
            r += 1
            for item in monthly:
                write_row(ws1, r, [f"{item['월']}월", item.get("접수", 0), item.get("종결", 0)])
                r += 1

    # 컬럼 너비
    for col, w in zip("ABCDEF", [16, 12, 12, 12, 12, 14]):
        ws1.column_dimensions[col].width = w

    # ════════════════════════════════════════════
    # 시트2: 사건 목록
    # ════════════════════════════════════════════
    ws2 = wb.create_sheet("사건목록")
    ws2.merge_cells("A1:G1")
    t2 = ws2["A1"]
    t2.value     = f"{period_label} 사건 목록"
    t2.font      = Font(bold=True, size=13)
    t2.alignment = CENTER
    ws2.row_dimensions[1].height = 24

    write_header(ws2, 2, ["접수번호", "신청인", "피신청인", "분쟁유형", "결과", "접수일자", "지역"])
    for i, c in enumerate(period_cases, 3):
        d = dict(c)
        write_row(ws2, i, [
            d.get("접수번호", ""), d.get("신청인_성명", ""),
            d.get("피신청인_성명", ""), d.get("분쟁유형", ""),
            d.get("결과", ""), d.get("접수일자", ""), d.get("지역", ""),
        ])

    for col, w in zip("ABCDEFG", [16, 12, 14, 12, 10, 12, 16]):
        ws2.column_dimensions[col].width = w

    # ── 다운로드 ─────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"보고서_{period_label.replace(' ','').replace('~','-')}_{date.today()}.xlsx"
    st.download_button(
        "⬇️ 다운로드",
        data=buf.read(),
        file_name=fname,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
