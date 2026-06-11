"""
접수대장 — Notion-style 테이블 + 우측 사이드 패널
행 클릭 → 사건 상세 패널 열림 / 문서 출력 · 엑셀 내보내기
"""
import streamlit as st
import pandas as pd
from datetime import date
from pathlib import Path
import sys

sys.path.insert(0, str(Path(__file__).parent.parent))
from core.db import init_db, get_all_cases, get_case, update_case
from core.status_resolver import resolve_status, STATUS_COLORS, CLOSED_STATUSES
from core.ui_styles import inject_css, page_header, status_badge, case_folder_path
from core.hwpx_handler import generate_hwpx
from core.excel_handler import generate_woopyeonmoa, generate_labeltek

if "db_initialized" not in st.session_state:
    init_db()
    st.session_state["db_initialized"] = True

inject_css()
page_header("📂", "접수대장", "사건 조회 · 문서 출력")

# 다른 페이지에서 진입 시 테이블 선택 초기화
if st.session_state.get("_current_page") != "접수대장":
    st.session_state["_tbl_key"] = st.session_state.get("_tbl_key", 0) + 1
st.session_state["_current_page"] = "접수대장"

STATUS_EMOJI = {
    "접수": "🔵", "회신대기": "🟡", "회신임박": "🟠", "회신지연": "🔴",
    "조정중지": "⚫", "불개시": "⚪", "개최예정": "🟢", "종결": "◻",
    "조정성립": "✅", "조정불성립": "❌",
}


# ══════════════════════════════════════════════
# 데이터 로드
# ══════════════════════════════════════════════
@st.cache_data(ttl=5)
def load_cases(year_sel, status_sel, type_sel, keyword):
    rows = get_all_cases(year=None if year_sel == "전체" else int(year_sel))
    data = []
    for r in rows:
        d = dict(r)
        d["진행상태"] = resolve_status(d)
        data.append(d)
    df = pd.DataFrame(data) if data else pd.DataFrame()
    if df.empty:
        return df
    if status_sel != "전체":
        df = df[df["진행상태"] == status_sel]
    if type_sel != "전체":
        df = df[df["분쟁유형"] == type_sel]
    if keyword:
        kw = keyword.strip().lower()
        mask = (
            df["접수번호"].str.lower().str.contains(kw, na=False)
            | df["신청인_성명"].str.lower().str.contains(kw, na=False)
            | df["피신청인_성명"].str.lower().str.contains(kw, na=False)
            | df["건물명"].fillna("").str.lower().str.contains(kw, na=False)
            | df["지역"].str.lower().str.contains(kw, na=False)
        )
        df = df[mask]
    return df


# ══════════════════════════════════════════════
# 필터 바
# ══════════════════════════════════════════════
f1, f2, f3, f4, f5 = st.columns([2, 2, 2, 3, 1])
cur_year = date.today().year
year_opts = ["전체"] + [str(y) for y in range(2050, 2025, -1)]
type_opts = [
    "전체",
    "하자",
    "관리인·관리위원 선임·해임 및 관리단 구성·운영",
    "공용부분 보존·관리·변경",
    "관리비 징수·관리·사용",
    "규약 제정·개정",
    "재건축 관련 철거·비용분담·구분소유권 귀속",
    "소음·약취 등 공동생활",
    "대지·부속시설 보존·관리·변경",
    "전유부분 사용방법",
    "관리비 외 수입 징수·관리·사용",
    "관리위탁계약 등 관리단 체결 계약",
    "기타",
]

with f1:
    default_year_idx = year_opts.index(str(cur_year)) if str(cur_year) in year_opts else 0
    year_sel = st.selectbox("연도", year_opts, index=default_year_idx, label_visibility="collapsed")
with f2:
    status_sel = st.selectbox("상태", ["전체"] + list(STATUS_COLORS.keys()),
                              label_visibility="collapsed")
with f3:
    type_sel = st.selectbox("분쟁유형", type_opts, label_visibility="collapsed")
with f4:
    keyword = st.text_input("검색", placeholder="접수번호, 이름, 건물명...",
                            label_visibility="collapsed")
with f5:
    st.markdown('<div style="height:28px"></div>', unsafe_allow_html=True)
    if st.button("↻", use_container_width=True, help="새로고침"):
        st.cache_data.clear()
        st.rerun()

df = load_cases(year_sel, status_sel, type_sel, keyword)

# ══════════════════════════════════════════════
# KPI 요약 바
# ══════════════════════════════════════════════
if not df.empty:
    total   = len(df)
    closed  = len(df[df["진행상태"].isin(CLOSED_STATUSES)])
    overdue = len(df[df["진행상태"] == "회신지연"])
    urgent  = len(df[df["진행상태"] == "회신임박"])
    active  = total - closed
    st.markdown(
        f'<div style="display:flex;gap:10px;margin-bottom:12px;flex-wrap:wrap">'
        f'<div class="kpi-card" style="flex:1;min-width:90px;border-top-color:#0066CC">'
        f'<div class="kpi-value">{total}</div><div class="kpi-label">전체</div></div>'
        f'<div class="kpi-card" style="flex:1;min-width:90px;border-top-color:#059669">'
        f'<div class="kpi-value">{active}</div><div class="kpi-label">진행 중</div></div>'
        f'<div class="kpi-card" style="flex:1;min-width:90px;border-top-color:#94A3B8">'
        f'<div class="kpi-value">{closed}</div><div class="kpi-label">종결</div></div>'
        f'<div class="kpi-card" style="flex:1;min-width:90px;border-top-color:#F59E0B">'
        f'<div class="kpi-value">{urgent}</div><div class="kpi-label">회신임박</div></div>'
        f'<div class="kpi-card" style="flex:1;min-width:90px;border-top-color:#DC2626">'
        f'<div class="kpi-value">{overdue}</div><div class="kpi-label">회신지연</div></div>'
        f'</div>',
        unsafe_allow_html=True,
    )

if df.empty:
    st.markdown(
        '<div class="card"><div class="empty-state">'
        '<div class="empty-icon">🔍</div>'
        '<div class="empty-title">조건에 맞는 사건이 없습니다</div>'
        '<div class="empty-sub">필터 조건을 변경하거나 검색어를 확인해 보세요</div>'
        '</div></div>',
        unsafe_allow_html=True,
    )
    st.stop()


# ══════════════════════════════════════════════
# 테이블 (전체 너비)
# ══════════════════════════════════════════════
st.markdown(
    '<p style="font-size:12px;color:#94A3B8;margin-bottom:6px">'
    '클릭으로 사건을 선택합니다  (다중 선택: Ctrl+클릭 또는 Shift+클릭)</p>',
    unsafe_allow_html=True,
)

SHOW_COLS = ["접수번호", "지역", "건물명", "신청인_성명", "피신청인_성명",
             "분쟁유형", "접수일자", "회신기한", "진행상태"]
df_disp = df[SHOW_COLS].copy()
df_disp["진행상태"] = df_disp["진행상태"].map(lambda s: f"{STATUS_EMOJI.get(s,'•')} {s}")
df_disp = df_disp.reset_index(drop=True)

tbl_event = st.dataframe(
    df_disp,
    use_container_width=True,
    hide_index=True,
    height=min(46 + len(df_disp) * 30, 500),
    row_height=30,
    on_select="rerun",
    selection_mode="multi-row",
    column_config={
        "접수번호":      st.column_config.TextColumn("접수번호",  width=88),
        "지역":          st.column_config.TextColumn("지역",      width=62),
        "건물명":        st.column_config.TextColumn("건물명",    width=100),
        "신청인_성명":   st.column_config.TextColumn("신청인",    width=80),
        "피신청인_성명": st.column_config.TextColumn("피신청인",  width=110),
        "분쟁유형":      st.column_config.TextColumn("유형",      width=120),
        "접수일자":      st.column_config.DateColumn("접수일자",  width=72, format="YY-MM-DD"),
        "회신기한":      st.column_config.DateColumn("회신기한",  width=72, format="YY-MM-DD"),
        "진행상태":      st.column_config.TextColumn("상태",      width=76),
    },
    key=f"case_table_{st.session_state.get('_tbl_key', 0)}",
)

sel_indices = tbl_event.selection.rows if tbl_event and tbl_event.selection else []
checked = [df_disp.iloc[i]["접수번호"] for i in sel_indices if i < len(df_disp)]
selected_id = checked[0] if len(checked) == 1 else (checked[-1] if len(checked) > 1 else None)

# 선택 사건 바뀌면 이전 문서 버튼 초기화
if st.session_state.get("_hwpx_selected") != selected_id:
    for k in list(st.session_state.keys()):
        if k.startswith("_hwpx_out_"):
            del st.session_state[k]
    st.session_state["_hwpx_selected"] = selected_id

# ══════════════════════════════════════════════
# 하단 액션 영역 (선택된 사건)
# ══════════════════════════════════════════════

# ── 다중 선택 시: 우편모아/라벨텍 묶음 출력
if len(checked) >= 2:
    st.markdown('<div style="height:10px"></div>', unsafe_allow_html=True)
    st.markdown(
        f'<div style="padding:12px 18px;background:#EFF6FF;border:1.5px solid #BFDBFE;'
        f'border-radius:10px;margin-bottom:10px">'
        f'<span style="font-weight:700;color:#1D4ED8;font-size:14px">'
        f'☑ {len(checked)}건 선택됨</span>'
        f'<span style="color:#3B82F6;font-size:12px;margin-left:10px">'
        f'{", ".join(checked)}</span>'
        f'</div>',
        unsafe_allow_html=True,
    )
    _inc_multi = st.checkbox("신청인 포함", key="inc_applicant_multi",
                             help="체크 시 신청인도 수취인으로 추가됩니다")
    mb1, mb2, mb3, mb4 = st.columns([1, 1, 1, 1])
    with mb1:
        if st.button("📮 우편모아 엑셀 (묶음)", use_container_width=True, key="gen_wp_multi", type="primary"):
            try:
                rows = [dict(get_case(cid)) for cid in checked if get_case(cid)]
                out = generate_woopyeonmoa(rows, include_applicant=_inc_multi)
                with open(out, "rb") as f:
                    st.download_button(
                        f"⬇ 우편모아 ({len(rows)}건) 다운로드", f,
                        file_name=out.name,
                        mime="application/vnd.ms-excel",
                        key="dl_wp_multi",
                    )
            except Exception as e:
                st.error(str(e))
    with mb2:
        if st.button("🏷️ 라벨텍 엑셀 (묶음)", use_container_width=True, key="gen_lb_multi", type="primary"):
            try:
                rows = [dict(get_case(cid)) for cid in checked if get_case(cid)]
                out = generate_labeltek(rows, include_applicant=_inc_multi)
                with open(out, "rb") as f:
                    st.download_button(
                        f"⬇ 라벨텍 ({len(rows)}건) 다운로드", f,
                        file_name=out.name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="dl_lb_multi",
                    )
            except Exception as e:
                st.error(str(e))
    with mb3:
        multi_status = st.selectbox(
            "일괄 상태 변경",
            ["", "종결", "조정성립", "조정불성립", "조정중지", "불개시"],
            label_visibility="collapsed",
            placeholder="⚡ 일괄 상태 변경...",
            key="multi_status_sel",
        )
    with mb4:
        c_apply, c_clear = st.columns(2)
        with c_apply:
            if st.button("적용", key="multi_status_apply",
                         disabled=not multi_status, use_container_width=True, type="primary"):
                st.session_state["_confirm_multi_status"] = multi_status
        with c_clear:
            if st.button("✕ 해제", use_container_width=True, key="clear_multi"):
                st.session_state["_tbl_key"] = st.session_state.get("_tbl_key", 0) + 1
                st.rerun()

    if st.session_state.get("_confirm_multi_status"):
        target = st.session_state["_confirm_multi_status"]
        STATUS_FIELD_MAP = {
            "종결":       {"결과": "종결",       "종결일자": date.today().isoformat()},
            "조정성립":   {"결과": "조정성립",   "종결일자": date.today().isoformat(), "개최여부": "개최"},
            "조정불성립": {"결과": "조정불성립", "종결일자": date.today().isoformat(), "개최여부": "개최"},
            "조정중지":   {"조정동의여부": "부동의"},
            "불개시":     {"개최여부": "불개시"},
        }
        fields = STATUS_FIELD_MAP.get(target, {})
        st.warning(
            f"선택된 **{len(checked)}건** ({', '.join(checked)})을 "
            f"**{target}** 으로 변경합니다. 계속하시겠습니까?"
        )
        mc1, mc2, _ = st.columns([1, 1, 5])
        with mc1:
            if st.button("✅ 확인", key="confirm_multi_yes", type="primary", use_container_width=True):
                for cid in checked:
                    update_case(cid, fields)
                st.session_state.pop("_confirm_multi_status", None)
                st.session_state["_tbl_key"] = st.session_state.get("_tbl_key", 0) + 1
                st.cache_data.clear()
                st.toast(f"{len(checked)}건 → {target} 변경 완료", icon="✅")
                st.rerun()
        with mc2:
            if st.button("✕ 취소", key="confirm_multi_no", use_container_width=True):
                st.session_state.pop("_confirm_multi_status", None)
                st.rerun()

elif not selected_id:
    st.markdown(
        '<div style="margin-top:10px;padding:20px 24px;background:#F8FAFC;border:1px dashed #CBD5E1;'
        'border-radius:10px;text-align:center;color:#94A3B8;font-size:13px">'
        '☑ 위 목록에서 사건을 선택하면 수정·상세보기·문서 출력 버튼이 표시됩니다'
        '</div>',
        unsafe_allow_html=True,
    )
else:
    case = get_case(selected_id)
    if case is None:
        st.error(f"{selected_id} 사건을 찾을 수 없습니다.")
    else:
        case = dict(case)
        status = resolve_status(case)
        badge_html = status_badge(status)

        today_str = date.today().isoformat()
        dl = case.get("회신기한") or ""
        if dl and dl < today_str:
            dl_color = "#DC2626"
            dl_label = f"{dl} · 기한 초과"
        elif dl:
            remain = (date.fromisoformat(dl) - date.today()).days
            dl_color = "#C25700" if remain <= 3 else "#334155"
            dl_label = f"{dl} · D-{remain}"
        else:
            dl_color = "#94A3B8"
            dl_label = "미정"

        # ── 선택 사건 요약 헤더
        st.markdown('<div style="height:10px"></div>', unsafe_allow_html=True)
        st.markdown(
            f'<div class="card" style="padding:14px 20px">'
            f'<div style="display:flex;align-items:center;gap:12px;flex-wrap:wrap">'
            f'<span style="font-size:1rem;font-weight:700;color:#0F172A">{case["접수번호"]}</span>'
            f'{badge_html}'
            f'<span style="color:#64748B;font-size:13px">{case.get("건물명") or ""}  ·  {case["지역"]}</span>'
            f'<span style="color:#334155;font-size:13px">'
            f'<b>신청인</b> {case["신청인_성명"]} &nbsp;<b>피신청인</b> {case["피신청인_성명"]}</span>'
            f'<span style="font-size:13px;color:{dl_color};margin-left:auto">회신기한 {dl_label}</span>'
            f'</div>'
            f'</div>',
            unsafe_allow_html=True,
        )

        # ── 버튼 영역 (5열 2행 균일 그리드)
        def _gen_hwpx(label: str, template: str, prefix: str, btn_key: str):
            if st.button(label, use_container_width=True, key=btn_key):
                cdata = dict(get_case(selected_id))
                try:
                    out = generate_hwpx(template, cdata, f"{prefix}_{selected_id}.hwpx")
                    st.session_state[f"_hwpx_out_{btn_key}"] = str(out)
                except Exception as e:
                    st.error(str(e))

            saved = st.session_state.get(f"_hwpx_out_{btn_key}")
            if saved:
                out = Path(saved)
                if out.exists():
                    dl_col, folder_col = st.columns(2)
                    with dl_col:
                        with open(out, "rb") as f:
                            st.download_button("⬇ 다운로드", f, file_name=out.name,
                                               mime="application/octet-stream",
                                               key=f"dl_{btn_key}")
                    with folder_col:
                        if st.button("📂 폴더 열기", key=f"open_{btn_key}"):
                            import subprocess
                            subprocess.Popen(f'explorer "{out.parent}"')

        # 1행: 이동 3 + 신청인포함 체크박스
        c1, c2, c3, c4, c5 = st.columns(5)
        with c1:
            if st.button("✏️ 수정하기", use_container_width=True, type="primary", key="act_edit"):
                st.session_state["edit_case"] = selected_id
                st.session_state["_edit_origin"] = "pages/3_접수대장.py"
                for k in list(st.session_state.keys()):
                    if k.startswith("_form_ready_") or k.startswith("inp_"):
                        del st.session_state[k]
                st.switch_page("pages/2_신규접수.py")
        with c2:
            if st.button("🔎 상세보기", use_container_width=True, key="act_detail"):
                st.session_state["detail_case"] = selected_id
                st.switch_page("pages/4_사건상세.py")
        with c3:
            if st.button("🗂️ 사건자료 폴더", use_container_width=True, key="act_folder"):
                import subprocess
                _root = Path(__file__).parent.parent / "output" / "사건자료"
                folder = case_folder_path(
                    _root, selected_id,
                    case.get("신청인_성명", ""), case.get("피신청인_성명", ""),
                )
                folder.mkdir(parents=True, exist_ok=True)
                subprocess.Popen(f'explorer "{folder}"')
        with c4:
            st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)
            _inc_single = st.checkbox("신청인 포함", key="inc_applicant_single")

        # 2행: 공문 3 + 엑셀 2
        d1, d2, d3, d4, d5 = st.columns(5)
        with d1:
            _gen_hwpx("📄 통지 공문", "1. 피신청인_통지_공문.hwpx", "피신청인_통지공문", "gen_notice")
        with d2:
            _gen_hwpx("🚫 조정중지 공문", "2. 조정중지 공문.hwpx", "조정중지_공문", "gen_stop")
        with d3:
            _gen_hwpx("📋 조정중지 통보서", "3. 조정중지 통보서.hwpx", "조정중지_통보서", "gen_stop2")
        with d4:
            if st.button("📮 우편모아", use_container_width=True, key="gen_wp"):
                try:
                    out = generate_woopyeonmoa([dict(get_case(selected_id))],
                                              include_applicant=_inc_single)
                    with open(out, "rb") as f:
                        st.download_button("⬇ 우편모아", f, file_name=out.name,
                                           mime="application/vnd.ms-excel", key="dl_wp")
                except Exception as e:
                    st.error(str(e))
        with d5:
            if st.button("🏷️ 라벨텍", use_container_width=True, key="gen_lb"):
                try:
                    out = generate_labeltek([dict(get_case(selected_id))],
                                           include_applicant=_inc_single)
                    with open(out, "rb") as f:
                        st.download_button("⬇ 라벨텍", f, file_name=out.name,
                                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                           key="dl_lb")
                except Exception as e:
                    st.error(str(e))

        # 3행: 나머지 공문 2종
        e1, e2, _, _, _ = st.columns(5)
        with e1:
            _gen_hwpx("🚫 불개시 공문", "5. 조정불개시_ 공문.hwpx", "조정불개시_공문", "gen_rej")
        with e2:
            _gen_hwpx("📋 불개시 통보서", "4. 조정불개시_통보서.hwpx", "조정불개시_통보서", "gen_rej2")

        # ── 빠른 상태 변경
        st.markdown('<hr style="margin:14px 0 10px">', unsafe_allow_html=True)
        qs1, qs2, qs3 = st.columns([3, 2, 1])
        with qs1:
            quick_status = st.selectbox(
                "빠른 상태 변경",
                ["", "종결", "조정성립", "조정불성립", "조정중지", "불개시"],
                label_visibility="collapsed",
                placeholder="⚡ 빠른 상태 변경...",
                key="quick_status_sel",
            )
        with qs2:
            st.markdown(
                f'<div style="padding-top:6px;font-size:12px;color:#64748B">'
                f'현재: <b>{status}</b></div>',
                unsafe_allow_html=True,
            )
        with qs3:
            apply_clicked = st.button(
                "적용", key="quick_status_apply",
                disabled=not quick_status, use_container_width=True,
            )

        if apply_clicked and quick_status:
            st.session_state["_confirm_status_change"] = quick_status

        if st.session_state.get("_confirm_status_change"):
            target = st.session_state["_confirm_status_change"]
            STATUS_FIELD_MAP = {
                "종결":       {"결과": "종결",       "종결일자": date.today().isoformat()},
                "조정성립":   {"결과": "조정성립",   "종결일자": date.today().isoformat(), "개최여부": "개최"},
                "조정불성립": {"결과": "조정불성립", "종결일자": date.today().isoformat(), "개최여부": "개최"},
                "조정중지":   {"조정동의여부": "부동의"},
                "불개시":     {"개최여부": "불개시"},
            }
            fields = STATUS_FIELD_MAP.get(target, {})
            st.warning(f"**{selected_id}** 상태를 **{target}** 으로 변경합니다. 계속하시겠습니까?")
            cc1, cc2, _ = st.columns([1, 1, 4])
            with cc1:
                if st.button("✅ 확인", key="confirm_qs_yes", type="primary", use_container_width=True):
                    update_case(selected_id, fields)
                    st.session_state.pop("_confirm_status_change", None)
                    st.cache_data.clear()
                    st.toast(f"{selected_id} → {target} 변경 완료", icon="✅")
                    st.rerun()
            with cc2:
                if st.button("✕ 취소", key="confirm_qs_no", use_container_width=True):
                    st.session_state.pop("_confirm_status_change", None)
                    st.rerun()


# ══════════════════════════════════════════════
# 엑셀 내보내기
# ══════════════════════════════════════════════
with st.expander("📤  엑셀 내보내기"):
    scope_col, _ = st.columns([3, 2])
    with scope_col:
        export_scope = st.radio(
            "범위",
            ["현재 필터 결과", "전체 (연도 무관)"],
            horizontal=True,
            label_visibility="collapsed",
        )

    if st.button("📥 엑셀 다운로드", type="primary"):
        import io, openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        if export_scope == "현재 필터 결과":
            full_rows = [dict(get_case(no)) for no in df["접수번호"] if get_case(no)] if not df.empty else []
        else:
            full_rows = [dict(r) for r in get_all_cases()]

        if not full_rows:
            st.warning("내보낼 데이터가 없습니다.")
        else:
            COLS = [
                ("접수번호", "접수번호"), ("접수연도", "접수연도"), ("접수일자", "접수일자"),
                ("지역", "지역"), ("분쟁유형", "분쟁유형"), ("건물명", "건물명"),
                ("건물소재지", "건물소재지"), ("신청인_성명", "신청인 성명"),
                ("신청인_지위", "신청인 지위"), ("신청인_주소", "신청인 주소"),
                ("신청인_우편번호", "신청인 우편번호"), ("신청인_연락처", "신청인 연락처"),
                ("피신청인_성명", "피신청인 성명"), ("피신청인_지위", "피신청인 지위"),
                ("피신청인_주소", "피신청인 주소"), ("피신청인_우편번호", "피신청인 우편번호"),
                ("피신청인_연락처", "피신청인 연락처"), ("안내도달일", "안내도달일"),
                ("회신기한", "회신기한"), ("회신접수일", "회신접수일"),
                ("조정동의여부", "조정동의여부"), ("개최여부", "개최여부"),
                ("결과", "결과"), ("종결일자", "종결일자"),
            ]
            thin = Side(style="thin")
            BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
            CENTER = Alignment(horizontal="center", vertical="center")

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "접수대장"

            for c_idx, (_, label) in enumerate(COLS, 1):
                cell = ws.cell(row=1, column=c_idx, value=label)
                cell.fill = PatternFill("solid", fgColor="0066CC")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = CENTER
                cell.border = BORDER

            for r_idx, row in enumerate(full_rows, 2):
                for c_idx, (field, _) in enumerate(COLS, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=row.get(field, "") or "")
                    cell.alignment = Alignment(vertical="center")
                    cell.border = BORDER

            for c_idx, (_, label) in enumerate(COLS, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = max(len(label) + 3, 12)
            ws.freeze_panes = "A2"

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            st.success(f"{len(full_rows)}건 준비 완료")
            st.download_button(
                "⬇️ 다운로드",
                data=buf.read(),
                file_name=f"접수대장_{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
